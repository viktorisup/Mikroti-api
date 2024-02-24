from dotenv import load_dotenv
import os
import routeros_api
import openpyxl
from openpyxl.styles import PatternFill, Font


load_dotenv()
user = os.environ.get('user')
passwd = os.environ.get('passwd')
ip_del = os.environ.get('ip_del')
ip_kus = os.environ.get('ip_kus')
ip_stom = os.environ.get('ip_stom')
ip_krut = os.environ.get('ip_krut')
ip_mos = os.environ.get('ip_mos')
ip_home = os.environ.get('ip_home')

routers_dict = {
    'Дом':
        {'ip': ip_home, 'user': user, 'passwd': passwd},
    # 'Делегатская и КДК' :
    #     {'ip': ip_del, 'user': user, 'passwd': passwd},
    # 'Кусково':
    #     {'ip': ip_kus, 'user': user, 'passwd': passwd},
    # 'Вучетича':
    #     {'ip': ip_stom, 'user': user, 'passwd': passwd},
    # 'Крутитский':
    #     {'ip': ip_krut, 'user': user, 'passwd': passwd},
    # 'Москва Сити':
    #     {'ip': ip_mos, 'user': user, 'passwd': passwd}
}


def create_newbook():
    newbook = openpyxl.Workbook()
    return newbook


def save_book(namebook, objbook):
    namebook = str(namebook + '.xlsx')
    objbook.save(namebook)


addr_book = create_newbook()
addr_book.remove(addr_book['Sheet'])
for i in routers_dict:
    addr_book.create_sheet(i)
    addr_book[i].column_dimensions['A'].width = 18
# sheet1 = addr_book['Делегатская']

    addr_book[i].column_dimensions['A'].width = 18
    addr_book[i]['A1'].font = Font(bold=True)
    addr_book[i].column_dimensions['B'].width = 18
    addr_book[i]['B1'].font = Font(bold=True)
    addr_book[i].column_dimensions['C'].width = 20
    addr_book[i]['C1'].font = Font(bold=True)
    addr_book[i].column_dimensions['D'].width = 20
    addr_book[i]['D1'].font = Font(bold=True)
    addr_book[i].column_dimensions['E'].width = 10
    addr_book[i]['E1'].font = Font(bold=True)
    addr_book[i].column_dimensions['G'].width = 10
    addr_book[i]['G1'].font = Font(bold=True)
    addr_book[i].column_dimensions['H'].width = 20
    addr_book[i]['H1'].font = Font(bold=True)
    addr_book[i].column_dimensions['I'].width = 20
    addr_book[i]['I1'].font = Font(bold=True)
    addr_book[i].column_dimensions['J'].width = 10
    addr_book[i]['J1'].font = Font(bold=True)


router_dict_obj = {}
try:
    for i in routers_dict:
        connection = routeros_api.RouterOsApiPool(
            routers_dict[i]['ip'],
            username=routers_dict[i]['user'],
            password=routers_dict[i]['passwd'],
            plaintext_login=True)
        api = connection.get_api()
        get_list_addr = api.get_resource('ip/address')
        list_addr = get_list_addr.get()
        router_dict_obj[i] = {'ADDRESS': list_addr}
        get_list_vlan = api.get_resource('interface/vlan')
        list_vlan = get_list_vlan.get()
        router_dict_obj[i].update(VLAN=list_vlan)
except Exception as e:
    with open('log_file', 'a', encoding='utf-8') as file:
        file.write(str(e) + '\n')



for name in router_dict_obj:
    cnt1 = 2
    for value in router_dict_obj[name]['ADDRESS']:
        addr_book[name].cell(row=1, column=1).value = 'Address'
        addr_book[name].cell(row=cnt1, column=1).value = value['address']
        addr_book[name].cell(row=1, column=2).value = 'Network'
        addr_book[name].cell(row=cnt1, column=2).value = value['network']
        addr_book[name].cell(row=1, column=3).value = 'Interface'
        addr_book[name].cell(row=cnt1, column=3).value = value['interface']
        if 'comment' in value:
            addr_book[name].cell(row=1, column=4).value = 'Comment'
            addr_book[name].cell(row=cnt1, column=4).value = value['comment']
        addr_book[name].cell(row=1, column=5).value = 'Disabled'
        addr_book[name].cell(row=cnt1, column=5).value = value['disabled']
        addr_book[name]['F1'].fill = PatternFill('solid', fgColor="7FC7FF")
        addr_book[name]['F' + str(cnt1)].fill = PatternFill('solid', fgColor="7FC7FF")
        cnt1 += 1

for name in router_dict_obj:
    cnt1 = 2
    for value in router_dict_obj[name]['VLAN']:
        addr_book[name].cell(row=1, column=7).value = 'Vlan'
        addr_book[name].cell(row=cnt1, column=7).value = value['vlan-id']
        addr_book[name].cell(row=1, column=8).value = 'Name'
        addr_book[name].cell(row=cnt1, column=8).value = value['name']
        addr_book[name].cell(row=1, column=9).value = 'Interface'
        addr_book[name].cell(row=cnt1, column=9).value = value['interface']
        addr_book[name].cell(row=1, column=10).value = 'Disabled'
        addr_book[name].cell(row=cnt1, column=10).value = value['disabled']
        cnt1 += 1


save_book('addr2', addr_book)